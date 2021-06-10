#!/usr/bin/env python
# coding: utf-8

# 1. TDM에서 특정 문서에서만 자주 사용된 단어에 대해서는 바이럴 마케팅 단어로 판단한다. (업체 정보에 대한 과한 어필) #
#      ->  추후 머신러닝을 통해 해당 단어들을 학습하고, 바이럴 마케팅 구분 문자인지를 학습시킨다.
# 2. 고료에 관한 언급이 있는 경우 바이럴 마케팅 글로 판단한다. #
# 3. Co-occurrence에서 일반적으로 사용되지 않는 문자 조합이 있는 경우, 어색한 문장으로 판단하여 바이럴 마케팅 데이터로 의심한다.
#     -> 추후 머신러닝을 도입하여 일반적인 문자 조합에 대한 학습
# 4. TDM에서 바이럴 마케팅 글에 사용된 문자 중 1 및 2에 대한 문자를 머신러닝이 학습하고, 중요도를 부여하여 추후 다른 글을 분석할 때, 바이럴 마케팅 글 여부를 판단하는 척도로 사용한다.
# 5. 다른 글과 비교할 때 유사한 문장 수를 분석하여 바이럴 마케팅 글 구분에 대한 척도로 사용한다. #
# 6. 일반적으로 사용하지 않은 단어를 사용한 경우 바이럴 마케팅 문구로 판단. #
#     -> 추후 머신러닝을 통해 가중치 부여 및 판단 척도로 사용
# 7. Co-occurrence matrix 분석 결과와 머신러닝을 통해 바이럴 마케팅 글에서 사용한 문자 조합(문장)에 바이럴 마케팅 데이터로써의 가중치를 부여한다.
#     -> 추후 새로운 데이터가 들어왔을 때, 바이럴 마케팅 글을 판별하는데 사용한다.

# In[20]:


import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from IPython.core.display import display, HTML

display(HTML("<style>.container { width:100% !important; }</style>"))
pd.set_option('display.max_rows', 999) # pd.options.display.max_rows = 999
pd.set_option('display.max_columns', 999) # pd.options.display.max_columns = 999
pd.set_option('display.width', 1000)

# 한글폰트 적용
plt.rcParams['font.family'] = 'Malgun Gothic'


# In[21]:


absolute_file_address = input("")

# read text data set
text_set = pd.read_csv(absolute_file_address)
text_set = text_set.fillna("")
text_set


# In[22]:


text_set.shape


# In[23]:


text_set = text_set.astype({'text':'string', 'url':'string'})
text_set.dtypes


# In[ ]:


from sklearn.feature_extraction.text import CountVectorizer
from konlpy.tag import Okt

import nltk
from nltk import bigrams
import itertools

import copy
import openpyxl
import csv


# co-occurrence matrix를 형성하는 함수
def generate_co_occurrence_matrix(corpus):
    vocab = set(corpus)
    vocab = list(vocab)
    vocab_index = {word: i for i, word in enumerate(vocab)}

    bi_grams = list(bigrams(corpus))
    bigram_freq = nltk.FreqDist(bi_grams).most_common(len(bi_grams))

    co_occurrence_matrix = np.zeros((len(vocab), len(vocab)))

    for bigram in bigram_freq:
        current = bigram[0][1]
        previous = bigram[0][0]
        count = bigram[1]
        pos_current = vocab_index[current]
        pos_previous = vocab_index[previous]
        co_occurrence_matrix[pos_current][pos_previous] = count

    co_occurrence_matrix = np.matrix(co_occurrence_matrix, dtype=np.int32)
    return co_occurrence_matrix, vocab_index


f = open('submission.csv', 'w', newline='')
wr = csv.writer(f)
# 과도한 정보 표기 , 고유 단어 사용, 고료 표기, 유사한 문장 사용, 유사한 문장 사용
wr.writerow(["url", "tmi", "use_unique_words", "indication_reward", "use_similar_sentences(including_body)", "use_similar_sentences(without_body)"])

# wb = openpyxl.load_workbook('submission.csv')
# sheet = wb.active
    
# sheet.cell(row=1, column=1).value = "과도한 정보 표기"
# sheet.cell(row=1, column=2).value = "고유 단어 사용"
# sheet.cell(row=1, column=3).value = "고료 표기"
# sheet.cell(row=1, column=4).value = "유사한 문장 사용(본문 포함)"
# sheet.cell(row=1, column=5).value = "유사한 문장 사용(본문 미포함)"

# wb.save("submission.csv")

analysis_count = 0
while analysis_count < len(text_set):
    if analysis_count+5 < len(text_set)-1:
        text_sub_set = text_set[analysis_count : analysis_count+5]
    else:
        text_sub_set = text_set[analysis_count : len(text_set)-1]
        
    
    # matrix(DataFrame)에서 text데이터를 가져온다.
    text_list = text_sub_set['text']
    url_list = text_sub_set['url']
    url_list = url_list.tolist()
    
    # 한국어 조사 정리.
    # 한국어 문법론 대가 이익섭 서울대 명예교수님께서 쓰신 '한국어문법'참고.
    # 자연어처리 분야에서 한국어 조사는 분석의 까다로움으로 인해 전처리 때 아예 제거되는 불용어(stop_words)로 취급됨. 
    stop_words = ["은", "는", "을", "를", "이", "가", "의", "에", "로", "으로", "과", "와", "도", "에서", "만"
                 , "이나", "나", "까지", "부터", "에게", "보다", "께", "처럼", "이라도", "라도", "으로서", "로서"
                 , "조차", "만큼", "같이", "마저", "이나마", "나마", "한테", "더러", "에게서", "한테서", "께서"
                 , "이야", "이라야"]

    commissional_words = ["협찬", "고료", "광고", "후원", "원고"]

    # Open Korea Text를 사용한 명사 추출 모듈 형성 - tokenizer진행
    # 본래 CountVectorizer는 토크나이징과 벡터화를 동시해 해주나, 이는 한국어를 대상으로하지는 않음.
    # 따라서 Okt모듈을 통해 토크나이징을 먼저 한 후 CountVectorizer작업 필요
    okt=Okt()
    text_token_set = list()

    count = 0
    for text in text_list.tolist():
        text_token_set.append(okt.nouns(text))

    # 시스템이 2층, 2층에, 2층으로 와 같은 글들을 각기 다른 하나의 명사 단어로 판단할수 있기에 토크나이징을 처리함


    # 출력결과 NxM: N개의 데이터에서 M개의 데이터를 뽑아냄: 벡터화
    # 결과적으로 각 문서에 어떤 단어가 몇번 등장했는지를 파악할 수 있음
    try:
        cv = CountVectorizer(tokenizer=lambda x: x, lowercase=False)
        tdm = cv.fit_transform(text_token_set)
    except ValueError:
#         dummy_value = np.zeros(shape=(len(text_list),), dtype=np.int64)
        
#         wb = openpyxl.load_workbook('submission.csv')
#         sheet = wb.active

#         for row in range(1, 6):
#             sheet.cell(row=analysis_count+row+1, column=1).value = dummy_value[row-1]
#             sheet.cell(row=analysis_count+row+1, column=2).value = dummy_value[row-1]
#             sheet.cell(row=analysis_count+row+1, column=3).value = dummy_value[row-1]
#             sheet.cell(row=analysis_count+row+1, column=4).value = dummy_value[row-1]
#             sheet.cell(row=analysis_count+row+1, column=5).value = dummy_value[row-1]

#         wb.save("submission.csv")

        for row in range(0, 5):
            wr.writerow([0, 0, 0, 0, 0])

        analysis_count = analysis_count+5


    # TDM 출력
    # TDM 대규모 데이터에서 대부분의 값은 0으로 나타낼 것임.
    # 이유는 하나의 텍스트에 2000개 종류의 단어를 사용한다 해도, 전체 단어 셋은 몇 만개는 될 것이기에
    # 메모리 부족 문제를 초래하기 쉽기 때문에, CountVectorizer는 희소행렬을 사용하기 때문.
    # 다음 출력 결과는 희소행렬임
    tdm_dataframe = pd.DataFrame(tdm.toarray())

    # TDM 분석: 너무 독자적인 단어 사용, 너무 TMI적 단어 사용 감지 (평균 * 5회 보다 사용수가 많거나 같고, 5회 이상 사용된 단어)
    use_tmi_words_value = np.zeros(shape=(len(text_list),), dtype=np.int64)
    # TDM 분석: 모든 독자적인 단어에 대한 수치 (혼자 사용된 단어가 10개중 1개 미만 일 경우)
    own_words_value = np.zeros(shape=(len(text_list),), dtype=np.int64)

    for col in range(0, tdm_dataframe.shape[1]):
        avg = sum(tdm_dataframe[col], 0.0) / len(tdm_dataframe[col]) # avg * 5 < 특정한 문서에서만 TMI적으로 사용된 단어 사용 빈도

        for row in range(0, len(tdm_dataframe[col])):
            if tdm_dataframe[col][row] >= (avg * 5) and tdm_dataframe[col][row] >= 5:
                use_tmi_words_value[row] = use_tmi_words_value[row] + 1

    print("use_tmi_words_value:", use_tmi_words_value)


    # TDM 분석: 모든 독자적인 단어에 대한 수치 (혼자 사용된 단어가 10개중 1개 미만 일 경우)
    own_words_value = np.zeros(shape=(len(text_list),), dtype=np.int64)

    for col in range(0, tdm_dataframe.shape[1]):
        used_value = 0
        for row in tdm_dataframe[col]:
            if row != 0:
                used_value += 1
        for row in range(0, len(tdm_dataframe[col])):
            if(len(text_list) == 0):
                continue
            
            if len(text_list) >= 20:
                if (used_value / 20) <= 1 and tdm_dataframe[col][row] != 0: 
                    own_words_value[row] = own_words_value[row] + 1

            else:
                if(used_value / len(text_list)) <= 1 and tdm_dataframe[col][row] != 0:
                    own_words_value[row] = own_words_value[row] + 1


    print("own_words_value", own_words_value)


    commissional_words_value = np.zeros(shape=(len(text_list),), dtype=np.int64)

    for index in range(0, len(text_list)):
        for word in commissional_words:
            if(len(text_list) == 0):
                continue
            if(text_list.tolist()[index].find(word) != -1):
                commissional_words_value[index] = 1

    print("commissional_words_value:", commissional_words_value)


    # # Co-occurrence matrix 형성


    co_occurrence_matrix_list = list()

    for text in text_list:
        tmp = text.split("\n")

        text_data = [okt.nouns(line) for line in tmp]
        text_data = list(itertools.chain.from_iterable(text_data))

        matrix, vocab_index = generate_co_occurrence_matrix(text_data)
        matrix_dataframe = pd.DataFrame(matrix, index=vocab_index, columns=vocab_index)
        co_occurrence_matrix_list.append(matrix_dataframe)
        

    # Co-occurrence matrix 분석: 너무 독자적인 단어 사용, 너무 TMI적 단어 사용 감지
    awkward_sentence_value = np.zeros(shape=(len(text_list),), dtype=np.int64)


    # # 비슷한 문장 판독

    # 1. 모든 문장에 대해 BoW적용
    # 2. 모든 text 대이터에 대해 대해 유사글 수치 부여 (갱신)
    # 3. 모든 text 데이터에 대해 유사글 수치 평균 계산


    # text데이터 별 비슷한 문장 수 리스트
    similar_sentence_value_all = np.zeros(shape=(len(text_list),), dtype=np.int64)
    similar_sentence_value_bes = np.zeros(shape=(len(text_list),), dtype=np.int64)

    # text_list토큰화
    text_data_token_set = list()
    for text in text_list:
        lines = text.split('\n')
        lines_token_set = [okt.nouns(line) for line in lines]

        text_data_token_set.append(lines_token_set)

    comp_list = list(itertools.chain.from_iterable(text_data_token_set))

    for i in range(0, len(text_list)):
        for line in text_data_token_set[i]:
            for comp in comp_list:
                line_len = len(line)
                comp_len = len(comp)
                line_to_comp_sub_len = len([x for x in line if x not in comp])
                comp_to_line_sub_len = len([x for x in comp if x not in line])

                if line_len + comp_len != 0:
                    if (line_to_comp_sub_len + comp_to_line_sub_len) / (line_len + comp_len) < 0.5:
                        similar_sentence_value_all[i] = similar_sentence_value_all[i] + 1

    for i in range(0, len(text_list)):
        for line in text_data_token_set[i]:

            explore_dest = copy.deepcopy(text_data_token_set)
            del explore_dest[i]
            for comp in list(itertools.chain.from_iterable(explore_dest)):
                line_len = len(line)
                comp_len = len(comp)
                line_to_comp_sub_len = len([x for x in line if x not in comp])
                comp_to_line_sub_len = len([x for x in comp if x not in line])

                if line_len + comp_len != 0:
                    if (line_to_comp_sub_len + comp_to_line_sub_len) / (line_len + comp_len) < 0.5:
                        similar_sentence_value_bes[i] = similar_sentence_value_bes[i] + 1

    print("similar_sentence_value_all:", similar_sentence_value_all)
    print("similar_sentence_value_bes: ", similar_sentence_value_bes)
    print("Analysis position:", analysis_count, "->", analysis_count+5)
    print("\n")
    
    
#     wb = openpyxl.load_workbook('submission.csv')
#     sheet = wb.active

#     for row in range(1, 6):
#         if(len(use_tmi_words_value) < row):
#             break
#         sheet.cell(row=analysis_count+row+1, column=1).value = use_tmi_words_value[row-1]
#         sheet.cell(row=analysis_count+row+1, column=2).value = own_words_value[row-1]
#         sheet.cell(row=analysis_count+row+1, column=3).value = commissional_words_value[row-1]
#         sheet.cell(row=analysis_count+row+1, column=4).value = similar_sentence_value_all[row-1]
#         sheet.cell(row=analysis_count+row+1, column=5).value = similar_sentence_value_bes[row-1]
        
#     wb.save("submission.csv")


    for row in range(0, 5):
        if(len(use_tmi_words_value) <= row):
            break
        wr.writerow([url_list[row], use_tmi_words_value[row], own_words_value[row], commissional_words_value[row], similar_sentence_value_all[row], similar_sentence_value_bes[row]])
        
        
    analysis_count = analysis_count+5
    
f.close()


# # 결과 기록

# In[ ]:


print("text_module_finish")


# In[ ]:




